"""
export_courses_by_type.py
─────────────────────────
Fetches courses from the /api/v1/courses/by-type endpoint and writes
the results to a formatted Excel (.xlsx) file, grouped by program.

Usage:
    python export_courses_by_type.py --course-type-id 1
    python export_courses_by_type.py --course-type-id 2 --base-url http://localhost:5000 --output my_report.xlsx

Requirements:
    pip install requests openpyxl
"""

import argparse
import sys
from datetime import datetime
from pathlib import Path

try:
    import requests
except ImportError:
    sys.exit("Missing dependency: run  pip install requests")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing dependency: run  pip install openpyxl")


# ─── Helpers ──────────────────────────────────────────────────────────────────

def fetch_courses(base_url: str, course_type_id: int) -> list:
def fetch_courses(base_url: str, course_type_id: int, verify: bool = True, proxies: dict = None) -> list:
    """Call the API and return the list of program groups."""
    url = f"{base_url.rstrip('/')}/api/v1/courses/by-type"
    # Trust system proxy environment variables (HTTP_PROXY, HTTPS_PROXY) automatically.
    # An explicit proxies dict overrides them; None means let requests read the env vars.
    session = requests.Session()
    if proxies:
        session.proxies.update(proxies)

    try:
        response = requests.get(url, params={"course_type_id": course_type_id}, timeout=10)
        response = session.get(
            url,
            params={"course_type_id": course_type_id},
            timeout=15,
            verify=verify,
        )
        response.raise_for_status()
    except requests.exceptions.ConnectionError:
        sys.exit(f"Could not connect to the server at {base_url}. Is the Flask dev server running?")
    except requests.exceptions.HTTPError as e:
    except requests.exceptions.SSLError as e:
        sys.exit(
            f"SSL certificate error: {e}\n"
            "  Tip: if the server uses a self-signed cert, re-run with --no-verify"
        )
    except requests.exceptions.ConnectionError as e:
        sys.exit(
            f"Could not connect to {base_url}\n"
            f"  Detail: {e}\n"
            "  Tips:\n"
            "    • Check that the URL scheme is correct (http vs https)\n"
            "    • If behind a university/corporate proxy, set the HTTPS_PROXY env var:\n"
            "        export HTTPS_PROXY=http://proxy.youruniversity.edu:8080\n"
            "    • Then re-run the script"
        )
    except requests.exceptions.Timeout:
        sys.exit(f"Request timed out after 15 seconds. The server at {base_url} may be slow or unreachable.")
    except requests.exceptions.HTTPError:
        sys.exit(f"API error {response.status_code}: {response.text}")
    except requests.exceptions.RequestException as e:
        sys.exit(f"Request failed: {e}")

    data = response.json()
    programs = data.get("programs", [])
    return programs


def build_excel(programs: list, course_type_id: int, output_path: Path) -> None:
    """Write the program/course data to a formatted Excel workbook."""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Courses by Type"

    # ── Style definitions ────────────────────────────────────────────────────

    HEADER_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
    PROGRAM_FILL  = PatternFill("solid", fgColor="D6E4F0")   # light blue
    ALT_ROW_FILL  = PatternFill("solid", fgColor="F2F2F2")   # light grey

    HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
    PROGRAM_FONT  = Font(bold=True, color="1F3864", size=10)
    BODY_FONT     = Font(size=10)

    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT   = Alignment(horizontal="left",   vertical="center")

    thin = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    COLUMNS = ["Program", "Course Code", "Course Title", "Unit"]
    COL_WIDTHS = [35, 16, 50, 8]

    # ── Title row ────────────────────────────────────────────────────────────

    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = f"Courses — Type ID {course_type_id}   |   Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    title_cell.font  = Font(bold=True, size=12, color="1F3864")
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 24

    # ── Column header row ────────────────────────────────────────────────────

    for col_idx, header in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = BORDER
    ws.row_dimensions[2].height = 18

    # ── Data rows ────────────────────────────────────────────────────────────

    current_row = 3
    total_courses = 0

    for program in programs:
        program_name = program.get("name", "Unknown Program")
        courses      = program.get("courses", [])

        if not courses:
            continue

        # Program sub-header row
        ws.merge_cells(f"A{current_row}:D{current_row}")
        prog_cell = ws.cell(row=current_row, column=1, value=f"  {program_name}")
        prog_cell.font      = PROGRAM_FONT
        prog_cell.fill      = PROGRAM_FILL
        prog_cell.alignment = LEFT
        prog_cell.border    = BORDER
        ws.row_dimensions[current_row].height = 16
        current_row += 1

        for i, course in enumerate(courses):
            row_fill = ALT_ROW_FILL if i % 2 == 1 else None

            values = [
                "",
                course.get("code", ""),
                course.get("title", ""),
                course.get("unit", ""),
            ]
            aligns = [CENTER, CENTER, LEFT, CENTER]

            for col_idx, (val, align) in enumerate(zip(values, aligns), start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font      = BODY_FONT
                cell.alignment = align
                cell.border    = BORDER
                if row_fill:
                    cell.fill = row_fill

            ws.row_dimensions[current_row].height = 15
            current_row += 1
            total_courses += 1

        # Blank spacer between programs
        current_row += 1

    # ── Column widths ────────────────────────────────────────────────────────

    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Freeze panes below headers ───────────────────────────────────────────

    ws.freeze_panes = "A3"

    # ── Summary footer ───────────────────────────────────────────────────────

    current_row += 1
    ws.merge_cells(f"A{current_row}:D{current_row}")
    summary_cell = ws.cell(
        row=current_row, column=1,
        value=f"Total: {len(programs)} program(s),  {total_courses} course(s)"
    )
    summary_cell.font      = Font(italic=True, size=9, color="666666")
    summary_cell.alignment = LEFT

    wb.save(output_path)


# ─── Entry point ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Export courses by course type from the API to an Excel file."
    )
    parser.add_argument(
        "--course-type-id",
        type=int,
        required=True,
        help="The course type ID to filter by (e.g. 1)"
    )
    parser.add_argument(
        "--base-url",
        default="http://localhost:5000",
        help="Base URL of the running Flask API (default: http://localhost:5000)"
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output Excel filename (default: courses_type_<id>_<timestamp>.xlsx)"
    )
    parser.add_argument(
        "--no-verify",
        action="store_true",
        default=False,
        help="Disable SSL certificate verification (use if the server has a self-signed cert)"
    )
    parser.add_argument(
        "--proxy",
        default=None,
        help="Explicit proxy URL, e.g. http://proxy.university.edu:8080  "
             "(by default the script also reads the HTTPS_PROXY / HTTP_PROXY env vars)"
    )

    args = parser.parse_args()

    proxies = {"http": args.proxy, "https": args.proxy} if args.proxy else None
    verify  = not args.no_verify

    output_filename = args.output or f"courses_type_{args.course_type_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = Path(output_filename)

    print(f"Fetching courses for course_type_id={args.course_type_id} from {args.base_url} ...")
    programs = fetch_courses(args.base_url, args.course_type_id)
    if not verify:
        print("  ⚠  SSL verification disabled.")
    if proxies:
        print(f"  Using proxy: {args.proxy}")

    programs = fetch_courses(args.base_url, args.course_type_id, verify=verify, proxies=proxies)

    if not programs:
        print("No courses found for the given course type ID. Excel file will not be created.")
        sys.exit(0)

    total = sum(len(p.get("courses", [])) for p in programs)
    print(f"  -> {len(programs)} program(s), {total} course(s) found.")

    print(f"Writing Excel file: {output_path} ...")
    build_excel(programs, args.course_type_id, output_path)

    print(f"Done! File saved: {output_path.resolve()}")


if __name__ == "__main__":
    main()
