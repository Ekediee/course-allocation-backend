"""
export_allocations_by_session.py
─────────────────────────────────
Fetches course allocations from /api/v1/allocation/by-session endpoint and writes
the results to a formatted Excel (.xlsx) file, grouped by semester and program.

Usage:
    python export_allocations_by_session.py --session-id 1
    python export_allocations_by_session.py --session-id 1 --semester-id 1
    python export_allocations_by_session.py --session-id 1 --base-url http://localhost:5000 --output my_allocations.xlsx

Requirements:
    pip install requests openpyxl
"""

import argparse
import sys
from datetime import datetime
from pathlib import Path

try:
    import requests
except ImportError:
    sys.exit("Missing dependency: run  pip install requests")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing dependency: run  pip install openpyxl")


def fetch_allocations(base_url: str, session_id: int, semester_id: int = None, verify: bool = True, proxies: dict = None) -> dict:
    """Call the API and return the allocation payload."""
    url = f"{base_url.rstrip('/')}/api/v1/allocation/by-session"
    params = {"session_id": session_id}
    if semester_id is not None:
        params["semester_id"] = semester_id

    session = requests.Session()
    if proxies:
        session.proxies.update(proxies)

    try:
        response = session.get(
            url,
            params=params,
            timeout=15,
            verify=verify,
        )
        response.raise_for_status()
    except requests.exceptions.SSLError as e:
        sys.exit(
            f"SSL certificate error: {e}\n"
            "  Tip: if the server uses a self-signed cert, re-run with --no-verify"
        )
    except requests.exceptions.ConnectionError as e:
        sys.exit(
            f"Could not connect to {base_url}\n"
            f"  Detail: {e}\n"
            "  Tips:\n"
            "    • Check that the URL scheme is correct (http vs https)\n"
            "    • If behind a proxy, set the HTTPS_PROXY env var or use --proxy\n"
        )
    except requests.exceptions.Timeout:
        sys.exit(f"Request timed out after 15 seconds connecting to {base_url}.")
    except requests.exceptions.HTTPError:
        sys.exit(f"API error {response.status_code}: {response.text}")
    except requests.exceptions.RequestException as e:
        sys.exit(f"Request failed: {e}")

    return response.json()


def build_excel(data: dict, output_path: Path) -> None:
    """Write the allocation data into a styled Excel workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Course Allocations"

    # Style definitions
    HEADER_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
    SEMESTER_FILL = PatternFill("solid", fgColor="2E5B82")   # medium steel blue
    PROGRAM_FILL  = PatternFill("solid", fgColor="D6E4F0")   # soft light blue
    ALT_ROW_FILL  = PatternFill("solid", fgColor="F2F2F2")   # light grey

    HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
    SEMESTER_FONT = Font(bold=True, color="FFFFFF", size=11)
    PROGRAM_FONT  = Font(bold=True, color="1F3864", size=10)
    BODY_FONT     = Font(size=10)

    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT   = Alignment(horizontal="left",   vertical="center")

    thin = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    COLUMNS = ["Semester", "Program", "Course Code", "Course Title", "Unit", "Group", "Lecturer Name"]
    COL_WIDTHS = [20, 32, 14, 42, 8, 12, 30]

    session_info = data.get("session", {})
    session_name = session_info.get("name", f"ID {session_info.get('id', '')}")
    semesters = data.get("semesters", [])

    # Title row
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"Course Allocations — Session: {session_name}   |   Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    title_cell.font  = Font(bold=True, size=12, color="1F3864")
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 24

    # Column headers
    for col_idx, header in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = BORDER
    ws.row_dimensions[2].height = 18

    current_row = 3
    total_courses = 0
    total_programs = 0

    for semester in semesters:
        semester_name = semester.get("name", "Unknown Semester")
        programs = semester.get("programs", [])

        if not programs:
            continue

        # Semester header banner
        ws.merge_cells(f"A{current_row}:G{current_row}")
        sem_cell = ws.cell(row=current_row, column=1, value=f"  ▶ Semester: {semester_name}")
        sem_cell.font      = SEMESTER_FONT
        sem_cell.fill      = SEMESTER_FILL
        sem_cell.alignment = LEFT
        sem_cell.border    = BORDER
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        for program in programs:
            program_name = program.get("name", "Unknown Program")
            courses = program.get("courses", [])

            if not courses:
                continue

            total_programs += 1

            # Program sub-header row
            ws.merge_cells(f"A{current_row}:G{current_row}")
            prog_cell = ws.cell(row=current_row, column=1, value=f"    • Program: {program_name}")
            prog_cell.font      = PROGRAM_FONT
            prog_cell.fill      = PROGRAM_FILL
            prog_cell.alignment = LEFT
            prog_cell.border    = BORDER
            ws.row_dimensions[current_row].height = 16
            current_row += 1

            for i, course in enumerate(courses):
                row_fill = ALT_ROW_FILL if i % 2 == 1 else None
                lecturer = course.get("lecturer") or {}

                values = [
                    "",
                    "",
                    course.get("code", ""),
                    course.get("title", ""),
                    course.get("unit", ""),
                    course.get("group_name") or "-",
                    lecturer.get("name") or "Unallocated",
                ]
                aligns = [CENTER, LEFT, CENTER, LEFT, CENTER, CENTER, LEFT]

                for col_idx, (val, align) in enumerate(zip(values, aligns), start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.font      = BODY_FONT
                    cell.alignment = align
                    cell.border    = BORDER
                    if row_fill:
                        cell.fill = row_fill

                ws.row_dimensions[current_row].height = 15
                current_row += 1
                total_courses += 1

            # Spacer between programs
            current_row += 1

        # Spacer between semesters
        current_row += 1

    # Column widths
    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A3"

    # Summary footer
    current_row += 1
    ws.merge_cells(f"A{current_row}:G{current_row}")
    summary_cell = ws.cell(
        row=current_row, column=1,
        value=f"Total: {len(semesters)} semester(s),  {total_programs} program group(s),  {total_courses} course allocation(s)"
    )
    summary_cell.font      = Font(italic=True, size=9, color="666666")
    summary_cell.alignment = LEFT

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="Export course allocations from the API to an Excel file."
    )
    parser.add_argument(
        "--session-id",
        type=int,
        required=True,
        help="The academic session ID to fetch allocations for (e.g. 1)"
    )
    parser.add_argument(
        "--semester-id",
        type=int,
        default=None,
        help="Optional semester ID to filter by (e.g. 1)"
    )
    parser.add_argument(
        "--base-url",
        default="http://localhost:5000",
        help="Base URL of the running Flask API (default: http://localhost:5000)"
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output Excel filename (default: allocations_session_<id>_<timestamp>.xlsx)"
    )
    parser.add_argument(
        "--no-verify",
        action="store_true",
        default=False,
        help="Disable SSL certificate verification (use if the server has a self-signed cert)"
    )
    parser.add_argument(
        "--proxy",
        default=None,
        help="Explicit proxy URL, e.g. http://proxy.university.edu:8080"
    )

    args = parser.parse_args()

    proxies = {"http": args.proxy, "https": args.proxy} if args.proxy else None
    verify  = not args.no_verify

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    if args.semester_id:
        default_filename = f"allocations_session_{args.session_id}_semester_{args.semester_id}_{timestamp}.xlsx"
    else:
        default_filename = f"allocations_session_{args.session_id}_{timestamp}.xlsx"

    output_filename = args.output or default_filename
    output_path = Path(output_filename)

    print(f"Fetching allocations for session_id={args.session_id}" + (f", semester_id={args.semester_id}" if args.semester_id else "") + f" from {args.base_url} ...")
    if not verify:
        print("  ⚠  SSL verification disabled.")
    if proxies:
        print(f"  Using proxy: {args.proxy}")

    data = fetch_allocations(args.base_url, args.session_id, args.semester_id, verify=verify, proxies=proxies)

    semesters = data.get("semesters", [])
    if not semesters:
        print("No allocations found for the given criteria. Excel file will not be created.")
        sys.exit(0)

    total = sum(
        len(prog.get("courses", []))
        for sem in semesters
        for prog in sem.get("programs", [])
    )
    print(f"  -> {len(semesters)} semester(s), {total} course allocation(s) found.")

    print(f"Writing Excel file: {output_path} ...")
    build_excel(data, output_path)

    print(f"Done! File saved: {output_path.resolve()}")


if __name__ == "__main__":
    main()
